import streamlit as st
import pandas as pd
import bcrypt
import hashlib
import io
import logging
from supabase import create_client
from datetime import date, datetime
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import Optional, List, Any

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO INICIAL & LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

st.set_page_config(
    page_title="Prates Compras",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    .stApp {background: #0f1419; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif}

    .login-box {max-width: 360px; margin: 80px auto; padding: 30px; background: #161b22; border-radius: 12px; border: 1px solid #30363d; box-shadow: 0 4px 16px rgba(0,0,0,0.3)}
    .login-header {text-align: center; margin-bottom: 20px}
    .login-logo {width: 55px; height: 55px; border-radius: 50%; background: linear-gradient(135deg, #238636, #2ea043); display: flex; align-items: center; justify-content: center; font-size: 22px; margin: 0 auto 10px; color: #fff}
    .login-title {font-size: 18px; font-weight: 700; color: #f0f6fc; margin: 5px 0}
    .login-sub {font-size: 11px; color: #8b949e}

    .stTextInput>div>div>input, .stNumberInput>div>div>input, .stSelectbox>div>div {background: #0d1117 !important; color: #e6edf3 !important; border: 1px solid #30363d !important; border-radius: 6px !important; padding: 8px 10px !important; font-size: 13px !important}

    .stButton>button {background: #238636 !important; color: #fff !important; border: none !important; border-radius: 6px !important; padding: 8px 16px !important; font-weight: 500 !important; font-size: 13px !important}
    .stButton>button:hover {background: #2ea043 !important}

    section[data-testid="stSidebar"] {background: #0d1117 !important; border-right: 1px solid #21262d !important; padding: 15px 10px !important}
    section[data-testid="stSidebar"] .stButton > button {
        text-align: left !important;
        justify-content: flex-start !important;
        width: 100% !important;
        padding: 8px 12px !important;
        display: flex !important;
        align-items: center !important;
    }
    section[data-testid="stSidebar"] .stButton > button p,
    section[data-testid="stSidebar"] .stButton > button span {
        text-align: left !important;
        margin: 0 !important;
    }
    .sidebar-header {text-align: center; padding: 15px 10px; margin-bottom: 15px; border-bottom: 1px solid #21262d}
    .sidebar-user {font-weight: 600; color: #f0f6fc; font-size: 13px; text-align: center}
    .sidebar-role {display: inline-block; background: #238636; color: #fff; font-size: 10px; padding: 2px 8px; border-radius: 10px; margin-top: 4px; font-weight: 600}

    .kpi-card {background: #161b22; border: 1px solid #21262d; border-radius: 10px; padding: 14px; text-align: center}
    .kpi-val {font-size: 22px; font-weight: 700; color: #f0f6fc; margin: 4px 0}
    .kpi-lbl {font-size: 11px; color: #8b949e; text-transform: uppercase; letter-spacing: 0.3px; font-weight: 600}

    .sec-hdr {background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 10px 14px; margin-bottom: 6px}
    .badge {display: inline-block; padding: 2px 7px; border-radius: 12px; font-size: 10px; font-weight: 600}
    .b-Pendente{background:rgba(210,153,34,.15);color:#D2991E}
    .b-Aprovado{background:rgba(88,166,255,.15);color:#58A6FF}
    .b-Comprado{background:rgba(163,113,247,.15);color:#A371F7}
    .b-Entregue{background:rgba(63,185,80,.15);color:#3FB950}
    .b-Cancelado{background:rgba(248,81,73,.15);color:#F85149}
    .b-Alta{background:rgba(248,81,73,.15);color:#F85149}
    .b-Media{background:rgba(210,153,34,.15);color:#D2991E}
    .b-Baixa{background:rgba(63,185,80,.15);color:#3FB950}

    .edit-panel {background: #161b22; border: 1px solid #30363d; border-radius: 10px; padding: 8px 12px; margin-bottom: 10px}
    .divider {height: 1px; background: #21262d; margin: 15px 0}
    .text-muted {color: #8b949e}
    .total-bar {background: #161b22; border: 1px solid #21262d; border-radius: 8px; padding: 10px 16px; margin-top: 10px; text-align: right}

    [data-testid="stPopover"]>div>button {background: #1C2128 !important; border: 1px solid #30363d !important; border-radius: 6px !important; color: #E6EDF3 !important; padding: 4px 8px !important; font-size: 12px !important}
    [data-testid="stPopoverBody"] {padding: 4px !important; min-width: 140px !important}
    [data-testid="stPopoverBody"] .stButton>button {font-size: 10px !important; padding: 4px 6px !important; min-height: 0 !important; border-radius: 4px !important}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
LOJAS = {
    "distribuidora": {"nome": "Prates Distribuidora", "cor": "#58A6FF", "icone": "📦"},
    "sublimacao":    {"nome": "Prates Sublimação",    "cor": "#3FB950", "icone": "🎨"}
}
STATUS_ALL = ["Pendente", "Aprovado", "Comprado", "Entregue", "Cancelado"]
STATUS_AT  = ("Pendente", "Aprovado")       # tuple → hashable para @st.cache_data
STATUS_HI  = ("Comprado", "Entregue", "Cancelado")
PRIO       = ["Alta", "Media", "Baixa"]
UNID       = ["UN", "CX", "PCT", "KG", "MT", "LT", "RL", "PAR"]

# Mapa de dias da semana em português
DIAS_PT = {0:"Segunda-feira", 1:"Terça-feira", 2:"Quarta-feira",
           3:"Quinta-feira",  4:"Sexta-feira", 5:"Sábado", 6:"Domingo"}
MESES_PT = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
            7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}

def data_pt(dt: datetime) -> str:
    """Retorna data formatada em português: Terça-feira, 14 de Abril de 2026 13:54"""
    dia_semana = DIAS_PT[dt.weekday()]
    mes        = MESES_PT[dt.month]
    return f"{dia_semana}, {dt.day:02d} de {mes} de {dt.year} {dt.hour:02d}:{dt.minute:02d}"

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def get_sb():
    try:
        return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])
    except Exception as e:
        logger.critical(f"DB Connection Failed: {e}")
        st.error("Erro de conexão com o banco. Verifique as credenciais.")
        st.stop()

sb = get_sb()

# ─────────────────────────────────────────────────────────────────────────────
# SEGURANÇA
# ─────────────────────────────────────────────────────────────────────────────
def hash_pwd(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_pwd(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except:
        return False

def check_session_timeout():
    """Verifica se a sessão expirou (30 minutos). CORRIGIDO: usa total_seconds()."""
    if "last_activity" in st.session_state:
        elapsed = (datetime.now() - st.session_state.last_activity).total_seconds()
        if elapsed > 1800:
            st.session_state.clear()
            st.warning("Sessão expirada. Faça login novamente.")
            st.rerun()
    st.session_state.last_activity = datetime.now()

# ─────────────────────────────────────────────────────────────────────────────
# CRUD — FORNECEDORES
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def get_fornecedores() -> List[dict]:
    try:
        return sb.table("pc_fornecedores").select("*").eq("ativo", True).order("nome").execute().data or []
    except Exception as e:
        logger.error(f"Erro fornecedores: {e}"); return []

def create_fornecedor(d: dict) -> bool:
    try:
        sb.table("pc_fornecedores").insert(d).execute()
        st.cache_data.clear(); return True
    except Exception as e:
        logger.error(f"Erro criar fornecedor: {e}"); return False

def update_fornecedor(fid: int, d: dict) -> bool:
    try:
        sb.table("pc_fornecedores").update(d).eq("id", fid).execute()
        st.cache_data.clear(); return True
    except Exception as e:
        logger.error(f"Erro atualizar fornecedor: {e}"); return False

def delete_fornecedor(fid: int) -> bool:
    try:
        sb.table("pc_fornecedores").update({"ativo": False}).eq("id", fid).execute()
        st.cache_data.clear(); return True
    except Exception as e:
        logger.error(f"Erro desativar fornecedor: {e}"); return False

# ─────────────────────────────────────────────────────────────────────────────
# CRUD — SEÇÕES
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=120)
def get_secoes(loja: str) -> List[dict]:
    try:
        return sb.table("pc_secoes").select("*").eq("loja", loja).eq("ativa", True).order("ordem").execute().data or []
    except Exception as e:
        logger.error(f"Erro seções: {e}"); return []

def create_secao(loja: str, nome: str) -> bool:
    try:
        ss = get_secoes(loja)
        ordem = (max(s["ordem"] for s in ss) + 1) if ss else 1
        sb.table("pc_secoes").insert({"loja": loja, "nome": nome, "ordem": ordem, "ativa": True}).execute()
        st.cache_data.clear(); return True
    except Exception as e:
        logger.error(f"Erro criar seção: {e}"); return False

def update_secao(sid: int, nome: str) -> bool:
    try:
        sb.table("pc_secoes").update({"nome": nome}).eq("id", sid).execute()
        st.cache_data.clear(); return True
    except Exception as e:
        logger.error(f"Erro atualizar seção: {e}"); return False

def delete_secao(sid: int) -> bool:
    try:
        sb.table("pc_secoes").update({"ativa": False}).eq("id", sid).execute()
        st.cache_data.clear(); return True
    except Exception as e:
        logger.error(f"Erro desativar seção: {e}"); return False

# ─────────────────────────────────────────────────────────────────────────────
# CRUD — ITENS
# ─────────────────────────────────────────────────────────────────────────────
def _validate_item(data: dict) -> Optional[str]:
    if not data.get("produto") or len(data["produto"]) > 150:
        return "Nome do produto inválido."
    if data.get("qtd", 0) < 0:
        return "Quantidade não pode ser negativa."
    if data.get("preco_unit", 0) < 0:
        return "Preço não pode ser negativo."
    return None

@st.cache_data(ttl=60)
def get_itens_secao(sid: int, status_filter: Optional[tuple] = None) -> List[dict]:
    """
    CORRIGIDO: status_filter é tuple (hashable) para funcionar com @st.cache_data.
    """
    try:
        q = sb.table("pc_itens").select("*").eq("secao_id", sid)
        if status_filter:
            q = q.in_("status", list(status_filter))
        return q.order("criado_em").execute().data or []
    except Exception as e:
        logger.error(f"Erro ao buscar itens da seção {sid}: {e}")
        return []

def create_item(sid: int, d: dict, user: str) -> Optional[str]:
    err = _validate_item(d)
    if err: return err
    try:
        d.update({"secao_id": sid, "criado_por": user,
                  "criado_em": datetime.now().isoformat(),
                  "atualizado_em": datetime.now().isoformat()})
        sb.table("pc_itens").insert(d).execute()
        st.cache_data.clear(); return None
    except Exception as e:
        logger.error(f"Erro criar item: {e}"); return "Erro ao salvar."

def update_item(iid: int, d: dict) -> Optional[str]:
    err = _validate_item(d)
    if err: return err
    try:
        d["atualizado_em"] = datetime.now().isoformat()
        sb.table("pc_itens").update(d).eq("id", iid).execute()
        st.cache_data.clear(); return None
    except Exception as e:
        logger.error(f"Erro atualizar item: {e}"); return "Erro ao atualizar."

def delete_item(iid: int) -> bool:
    try:
        sb.table("pc_itens").delete().eq("id", iid).execute()
        st.cache_data.clear(); return True
    except Exception as e:
        logger.error(f"Erro deletar item: {e}"); return False

def batch_update_status(ids: List[int], new_status: str) -> bool:
    if not ids: return False
    try:
        for iid in ids:
            sb.table("pc_itens").update(
                {"status": new_status, "atualizado_em": datetime.now().isoformat()}
            ).eq("id", iid).execute()
        st.cache_data.clear(); return True
    except Exception as e:
        logger.error(f"Erro batch: {e}"); return False

# ─────────────────────────────────────────────────────────────────────────────
# CRUD — USUÁRIOS
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def get_usuarios() -> List[dict]:
    try:
        return sb.table("pc_usuarios").select("id,nome,email,acesso,ativo").order("nome").execute().data or []
    except Exception as e:
        logger.error(f"Erro usuários: {e}"); return []

def create_usuario(nome: str, email: str, senha: str, acesso: str) -> Optional[str]:
    try:
        sb.table("pc_usuarios").insert({
            "nome": nome, "email": email.lower(),
            "senha_hash": hash_pwd(senha), "acesso": acesso, "ativo": True
        }).execute()
        st.cache_data.clear(); return None
    except Exception as e:
        logger.error(f"Erro criar usuário: {e}"); return "Erro ao criar."

def update_usuario(uid: int, d: dict) -> Optional[str]:
    try:
        sb.table("pc_usuarios").update(d).eq("id", uid).execute()
        st.cache_data.clear(); return None
    except Exception as e:
        logger.error(f"Erro atualizar usuário: {e}"); return "Erro ao atualizar."

# ─────────────────────────────────────────────────────────────────────────────
# UTILS
# ─────────────────────────────────────────────────────────────────────────────
def fmt_brl(v: Any) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "R$ 0,00"

def parse_date_safe(dt_str: Any) -> Optional[date]:
    if not dt_str: return None
    try: return date.fromisoformat(str(dt_str))
    except:
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try: return datetime.strptime(str(dt_str), fmt).date()
            except: continue
    return None

def days_until(dt_str: Any) -> Optional[int]:
    d = parse_date_safe(dt_str)
    return (d - date.today()).days if d else None

def badge(txt: str) -> str:
    c = {"Media": "Media", "Média": "Media"}.get(txt, txt)
    return f"<span class='badge b-{c}'>{txt}</span>"

def check_perm(loja: str) -> bool:
    u = st.session_state.get("usuario")
    if not u: return False
    a = u["acesso"]
    if a in ("admin", "ambas"): return True
    if loja == "distribuidora": return a in ("distribuidora", "op_dist", "op_ambas")
    if loja == "sublimacao":    return a in ("sublimacao", "op_sub", "op_ambas")
    return False

def is_op() -> bool:
    u = st.session_state.get("usuario")
    return bool(u and u["acesso"] in ("op_dist", "op_sub", "op_ambas"))

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for k, v in [("usuario", None), ("pagina", "dashboard"), ("site_ok", False)]:
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────────────────
# PORTÃO SITE_CODE  ← IMPLEMENTADO (estava faltando)
# ─────────────────────────────────────────────────────────────────────────────
if not st.session_state.site_ok:
    _, c, _ = st.columns([1, 0.8, 1])
    with c:
        st.markdown("""
        <div class="login-box">
            <div class="login-header">
                <div class="login-logo">🔑</div>
                <div class="login-title">Acesso Restrito</div>
                <div class="login-sub">Digite o código de acesso do site</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        with st.form("site_form"):
            codigo = st.text_input("Código", type="password",
                                   placeholder="Digite o código (sem aspas)",
                                   label_visibility="collapsed")
            if st.form_submit_button("Acessar", use_container_width=True, type="primary"):
                if codigo == st.secrets.get("SITE_CODE", ""):
                    st.session_state.site_ok = True
                    st.rerun()
                else:
                    st.error("Código incorreto.", icon="🔒")
        st.markdown("<div style='text-align:center;margin-top:20px;color:#8b949e;font-size:11px'>© 2026 Prates</div>",
                    unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def do_login(email: str, senha: str) -> bool:
    email = email.strip().lower()
    if not email or not senha:
        st.warning("Preencha email e senha.", icon="⚠️"); return False
    try:
        r = sb.table("pc_usuarios").select("*").eq("email", email).eq("ativo", True).execute()
        if not r.data:
            logger.warning(f"Login falho: {email}")
            st.error("Email ou senha incorretos.", icon="🔒"); return False
        u = r.data[0]
        if u["senha_hash"].startswith("$2b$"):
            if not verify_pwd(senha, u["senha_hash"]):
                st.error("Email ou senha incorretos.", icon="🔒"); return False
        else:
            if hashlib.sha256(senha.encode()).hexdigest() != u["senha_hash"]:
                st.error("Email ou senha incorretos.", icon="🔒"); return False
            # Migra hash para bcrypt
            sb.table("pc_usuarios").update({"senha_hash": hash_pwd(senha)}).eq("id", u["id"]).execute()
        st.session_state.usuario = u
        logger.info(f"Login: {email}"); return True
    except Exception as e:
        logger.error(f"Erro login: {e}")
        st.error("Erro interno. Tente novamente.", icon="❌"); return False

if not st.session_state.usuario:
    _, c, _ = st.columns([1, 0.8, 1])
    with c:
        st.markdown("""
        <div class="login-box">
            <div class="login-header">
                <div class="login-logo">🛒</div>
                <div class="login-title">Prates Compras</div>
                <div class="login-sub">Sistema de Gestão</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        with st.form("login_form"):
            em = st.text_input("Email", placeholder="seu@email.com", label_visibility="collapsed")
            se = st.text_input("Senha", type="password", placeholder="Digite sua senha", label_visibility="collapsed")
            if st.form_submit_button("Entrar", use_container_width=True, type="primary"):
                if do_login(em, se): st.rerun()
        st.markdown("<div style='text-align:center;margin-top:20px;color:#8b949e;font-size:11px'>© 2026 Prates</div>",
                    unsafe_allow_html=True)
    st.stop()

# Verificar timeout — CORRIGIDO: usa total_seconds()
check_session_timeout()

u = st.session_state.usuario

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div class="sidebar-header">
        <div class="sidebar-user">👤 {u['nome']}</div>
        <div class="sidebar-role">{u['acesso'].upper()}</div>
    </div>
    """, unsafe_allow_html=True)

    nav = [
        ("📊 Dashboard",    "dashboard"),
        ("📦 Distribuidora","distribuidora"),
        ("🎨 Sublimação",   "sublimacao"),
        ("📅 Histórico",    "historico"),
        ("📥 Exportar",     "exportar"),
        ("🏭 Fornecedores", "fornecedores"),
    ]
    if u["acesso"] == "admin":
        nav.append(("⚙️ Admin", "admin"))

    for label, page in nav:
        if st.button(label, use_container_width=True, key=f"nav_{page}"):
            st.session_state.pagina = page
            st.rerun()

    st.divider()
    if st.button("🚪 Sair", use_container_width=True):
        st.session_state.usuario  = None
        st.session_state.site_ok  = False
        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# PÁGINAS
# ─────────────────────────────────────────────────────────────────────────────

def pagina_dashboard():
    # Data em português — CORRIGIDO
    agora = datetime.now()
    st.markdown(
        f"<div style='color:#8b949e;font-size:12px;margin-bottom:12px'>🕐 {data_pt(agora)}</div>",
        unsafe_allow_html=True
    )
    st.markdown("<h2 style='font-size:20px;font-weight:700;color:#f0f6fc;margin-bottom:10px'>📊 Dashboard</h2>",
                unsafe_allow_html=True)

    @st.cache_data(ttl=300, show_spinner=False)
    def load_all_items():
        try:
            return sb.table("pc_itens").select("*, pc_secoes(nome,loja)").execute().data or []
        except Exception as e:
            logger.error(f"Erro dashboard: {e}"); return []

    with st.spinner("Carregando..."):
        todos = load_all_items()

    if not todos:
        st.markdown("""
        <div style='text-align:center;padding:40px;background:#161b22;border-radius:10px;border:1px dashed #30363d'>
            <div style='font-size:36px'>📦</div>
            <div class='text-muted' style='margin-top:8px'>Nenhum produto lançado.</div>
        </div>""", unsafe_allow_html=True)
        return

    df = pd.DataFrame(todos)
    df["loja"]  = df["pc_secoes"].apply(lambda x: x["loja"] if x else "")
    df["secao"] = df["pc_secoes"].apply(lambda x: x["nome"] if x else "")
    df["total"] = pd.to_numeric(df.get("total", 0), errors="coerce").fillna(0)

    # KPIs linha 1
    cols = st.columns(5)
    metrics = [
        ("💰 Total Geral",    fmt_brl(df["total"].sum()),                                            "#f0f6fc"),
        ("📦 Distribuidora",  fmt_brl(df[df["loja"]=="distribuidora"]["total"].sum()),               "#58A6FF"),
        ("🎨 Sublimação",     fmt_brl(df[df["loja"]=="sublimacao"]["total"].sum()),                  "#3FB950"),
        ("📋 Total Itens",    str(len(df)),                                                          "#8b949e"),
        ("✅ Entregues",      str((df["status"]=="Entregue").sum()),                                 "#3FB950"),
    ]
    for i, (lbl, val, cor) in enumerate(metrics):
        cols[i].markdown(
            f"<div class='kpi-card' style='border-top:3px solid {cor}'>"
            f"<div class='kpi-lbl'>{lbl}</div>"
            f"<div class='kpi-val' style='color:{cor}'>{val}</div></div>",
            unsafe_allow_html=True)

    # KPIs linha 2
    cols2 = st.columns(5)
    metrics2 = [
        ("🟡 Pendentes",  str((df["status"]=="Pendente").sum()),                     "#D2991E"),
        ("🔵 Aprovados",  str((df["status"]=="Aprovado").sum()),                     "#58A6FF"),
        ("🟣 Comprados",  str((df["status"]=="Comprado").sum()),                     "#A371F7"),
        ("❌ Cancelados", str((df["status"]=="Cancelado").sum()),                    "#F85149"),
        ("📊 Itens Ativos",str(len(df[df["status"].isin(list(STATUS_AT))])),         "#8b949e"),
    ]
    for i, (lbl, val, cor) in enumerate(metrics2):
        cols2[i].markdown(
            f"<div class='kpi-card' style='border-top:3px solid {cor}'>"
            f"<div class='kpi-lbl'>{lbl}</div>"
            f"<div class='kpi-val' style='color:{cor}'>{val}</div></div>",
            unsafe_allow_html=True)

    st.divider()

    # Gráficos
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("<div class='text-muted' style='font-size:12px;margin-bottom:8px'>POR STATUS</div>",
                    unsafe_allow_html=True)
        s = df.groupby("status").size().reset_index(name="qtd")
        fig = px.bar(s, x="status", y="qtd", color="status",
                     color_discrete_map={"Pendente":"#D2991E","Aprovado":"#58A6FF",
                                         "Comprado":"#A371F7","Entregue":"#3FB950","Cancelado":"#F85149"},
                     template="plotly_dark")
        fig.update_layout(showlegend=False, height=200,
                          margin=dict(t=10,b=10,l=10,r=10),
                          paper_bgcolor="#161b22", plot_bgcolor="#0d1117",
                          font=dict(color="#8b949e", size=10))
        st.plotly_chart(fig, use_container_width=True)

    with c2:
        st.markdown("<div class='text-muted' style='font-size:12px;margin-bottom:8px'>POR LOJA</div>",
                    unsafe_allow_html=True)
        l = df.groupby("loja").size().reset_index(name="qtd")
        l["nome"] = l["loja"].map({"distribuidora":"Distribuidora","sublimacao":"Sublimação"})
        fig2 = px.pie(l, names="nome", values="qtd",
                      color_discrete_sequence=["#58A6FF","#3FB950"], hole=0.5, template="plotly_dark")
        fig2.update_layout(height=200, margin=dict(t=10,b=10),
                           paper_bgcolor="#161b22", font=dict(color="#8b949e", size=10),
                           legend=dict(orientation="h", y=-0.1, font=dict(size=10)))
        st.plotly_chart(fig2, use_container_width=True)

    with c3:
        st.markdown("<div class='text-muted' style='font-size:12px;margin-bottom:8px'>TOP 5 PRODUTOS</div>",
                    unsafe_allow_html=True)
        top = df.groupby("produto")["qtd"].sum().sort_values(ascending=True).tail(5).reset_index()
        if not top.empty:
            fig3 = px.bar(top, x="qtd", y="produto", orientation="h",
                          template="plotly_dark", color_discrete_sequence=["#3FB950"])
            fig3.update_layout(showlegend=False, height=200,
                               margin=dict(t=10,b=10,l=10,r=10),
                               paper_bgcolor="#161b22", plot_bgcolor="#0d1117",
                               font=dict(color="#8b949e", size=10),
                               yaxis=dict(tickfont=dict(size=10)))
            st.plotly_chart(fig3, use_container_width=True)

    # Valor por Seção
    st.markdown("<div class='text-muted' style='font-size:12px;margin:10px 0 8px'>VALOR POR SEÇÃO</div>",
                unsafe_allow_html=True)
    sec = df.groupby(["secao","loja"])["total"].sum().reset_index().sort_values("total", ascending=True).tail(8)
    if not sec.empty:
        fig4 = px.bar(sec, x="total", y="secao", orientation="h", color="loja",
                      color_discrete_map={"distribuidora":"#58A6FF","sublimacao":"#3FB950"},
                      template="plotly_dark")
        fig4.update_layout(
            legend=dict(bgcolor="#161b22", orientation="h", yanchor="bottom", y=1.02, font=dict(size=10)),
            height=180, margin=dict(t=10,b=10,l=10,r=10),
            paper_bgcolor="#161b22", plot_bgcolor="#0d1117",
            font=dict(color="#8b949e", size=10), yaxis=dict(tickfont=dict(size=10)))
        st.plotly_chart(fig4, use_container_width=True)


def pagina_loja(loja: str):
    info = LOJAS[loja]
    st.markdown(f"<h2 style='font-size:18px;font-weight:700;color:#f0f6fc;margin-bottom:12px'>"
                f"{info['icone']} {info['nome']}</h2>", unsafe_allow_html=True)

    c1, c2, c3 = st.columns([3, 1, 1])
    busca = c1.text_input("", placeholder="Buscar produto, marca, SKU...",
                          label_visibility="collapsed", key=f"busc_{loja}")
    if c2.button("+ Produto", use_container_width=True, key=f"btnp_{loja}"):
        st.session_state[f"cp_{loja}"] = not st.session_state.get(f"cp_{loja}", False)
    if c3.button("Seções", use_container_width=True, key=f"btns_{loja}"):
        st.session_state[f"gs_{loja}"] = not st.session_state.get(f"gs_{loja}", False)

    f1, f2 = st.columns(2)
    fst = f1.radio("", ["Todos"]+list(STATUS_AT), horizontal=True,
                   key=f"fst_{loja}", label_visibility="collapsed")
    fpr = f2.radio("", ["Todas"]+PRIO, horizontal=True,
                   key=f"fpr_{loja}", label_visibility="collapsed")

    # Gerenciar Seções
    if st.session_state.get(f"gs_{loja}"):
        st.markdown("<div class='sec-hdr'><span style='font-weight:600;color:#58A6FF;font-size:13px'>"
                    "Gerenciar Seções</span></div>", unsafe_allow_html=True)
        with st.form(f"fns_{loja}"):
            n1, n2 = st.columns([4, 1])
            nn = n1.text_input("", placeholder="Nome da nova seção", label_visibility="collapsed")
            if n2.form_submit_button("+ Criar", type="primary", use_container_width=True):
                if nn.strip(): create_secao(loja, nn.strip()); st.rerun()
        for sec in get_secoes(loja):
            with st.form(f"fsec_{sec['id']}"):
                e1, e2, e3 = st.columns([4, 1, 1])
                v = e1.text_input("", value=sec["nome"], label_visibility="collapsed")
                if e2.form_submit_button("Salvar", type="primary", use_container_width=True):
                    update_secao(sec["id"], v.strip()); st.rerun()
                if e3.form_submit_button("Excluir", use_container_width=True):
                    delete_secao(sec["id"]); st.rerun()
        if st.button("Fechar", key=f"fgs_{loja}"):
            st.session_state[f"gs_{loja}"] = False; st.rerun()
        st.divider()

    # Novo Produto
    if st.session_state.get(f"cp_{loja}"):
        secs = get_secoes(loja)
        if not secs:
            st.warning("Crie uma seção primeiro.")
        else:
            st.markdown("<div class='sec-hdr' style='border-top:2px solid #238636'>"
                        "<span style='font-weight:600;color:#3FB950;font-size:13px'>Novo Produto</span>"
                        "</div>", unsafe_allow_html=True)
            fmc = {f["nome"]: f["id"] for f in get_fornecedores()}
            with st.form(f"fcp_{loja}"):
                r1 = st.columns([2, 2.5, 2, 1.5, 1.5, 2])
                sec_sel = r1[0].selectbox("Seção",      [s["nome"] for s in secs])
                prod    = r1[1].text_input("Produto *")
                marca   = r1[2].text_input("Marca")
                sku     = r1[3].text_input("SKU")
                ean     = r1[4].text_input("EAN")
                forn    = r1[5].selectbox("Fornecedor", ["(Nenhum)"]+list(fmc.keys()))

                r2 = st.columns([1.2, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1])
                qtd  = r2[0].number_input("Qtd",    min_value=0.0, step=1.0)
                un   = r2[1].selectbox("Unid",      UNID)
                prec = r2[2].number_input("Preço",  min_value=0.0, step=0.01, format="%.2f")
                prio = r2[3].selectbox("Prioridade",PRIO, index=1)
                dt   = r2[4].date_input("Necessidade", value=None, format="DD/MM/YYYY")
                img  = r2[5].text_input("URL Img",  placeholder="https://...")
                obs  = r2[6].text_input("Obs")
                r2[7].markdown("<br>", unsafe_allow_html=True)

                if r2[7].form_submit_button("Salvar", type="primary", use_container_width=True):
                    if prod.strip():
                        sid = next(s["id"] for s in secs if s["nome"] == sec_sel)
                        fid = fmc.get(forn) if forn != "(Nenhum)" else None
                        err = create_item(sid, {
                            "produto": prod.strip(), "marca": marca.strip(),
                            "sku": sku.strip(), "ean": ean.strip(),
                            "fornecedor_id": fid, "imagem_url": img.strip() or None,
                            "qtd": qtd, "unidade": un, "preco_unit": prec,
                            "total": round(qtd * prec, 2), "prioridade": prio,
                            "dt_necessidade": str(dt) if dt else None,
                            "obs": obs.strip(), "status": "Pendente"
                        }, u["nome"])
                        if err:
                            st.error(err)
                        else:
                            st.success("Salvo!")
                            st.session_state[f"cp_{loja}"] = False
                            st.rerun()
                    else:
                        st.warning("Informe o nome do produto.")
        st.divider()

    # Seções e Itens
    secoes     = get_secoes(loja)
    total_loja = 0.0
    fmc        = {f["id"]: f["nome"] for f in get_fornecedores()}
    sel_key    = f"sel_{loja}"
    if sel_key not in st.session_state:
        st.session_state[sel_key] = []
    marcados = st.session_state[sel_key]

    # Ações em lote
    if marcados and not is_op():
        st.markdown(
            f"<div style='background:#1C2128;border:1px solid #30363d;border-radius:6px;"
            f"padding:6px 10px;margin-bottom:10px'>"
            f"<b style='color:#58A6FF;font-size:12px'>{len(marcados)} selecionado(s)</b></div>",
            unsafe_allow_html=True)
        ba = st.columns(5)
        ops = [("Aprovar","Aprovado","primary"),("Comprado","Comprado","secondary"),
               ("Entregue","Entregue","secondary"),("Cancelar","Cancelado","secondary"),
               ("Limpar","",None)]
        for i, (lbl, sv, tp) in enumerate(ops):
            if tp and ba[i].button(lbl, key=f"lote_{sv}_{loja}", use_container_width=True, type=tp):
                batch_update_status(marcados, sv)
                st.session_state[sel_key] = []
                st.rerun()
            elif not tp and ba[i].button(lbl, key=f"lote_limpar_{loja}", use_container_width=True):
                st.session_state[sel_key] = []
                st.rerun()
        st.divider()

    for sec in secoes:
        itens_all = get_itens_secao(sec["id"], STATUS_AT)   # STATUS_AT já é tuple

        itens = itens_all[:]
        if fst != "Todos":  itens = [i for i in itens if i.get("status") == fst]
        if fpr != "Todas":  itens = [i for i in itens if i.get("prioridade") == fpr]
        if busca:
            b = busca.lower()
            itens = [i for i in itens if b in " ".join([
                i.get("produto",""), i.get("marca",""),
                i.get("sku",""), i.get("ean","")]).lower()]

        if (busca or fst != "Todos" or fpr != "Todas") and not itens:
            continue

        tsec  = sum(float(i.get("total") or 0) for i in itens_all)
        total_loja += tsec
        npend = sum(1 for i in itens_all if i.get("status") == "Pendente")

        titulo = f"📁 {sec['nome']} ({len(itens_all)} itens)"
        if npend > 0: titulo += f" · {npend} pendente(s)"

        with st.expander(titulo, expanded=True):
            if itens:
                for idx, item in enumerate(itens):
                    iid  = item["id"]
                    cols = st.columns([0.3, 3.5, 1.2, 0.8, 0.8, 1.2, 1, 1, 0.8])

                    # Checkbox seleção
                    sel = cols[0].checkbox("", value=iid in marcados,
                                           key=f"chk_{iid}_{idx}", label_visibility="collapsed")
                    if sel and iid not in marcados:
                        st.session_state[sel_key].append(iid)
                    elif not sel and iid in marcados:
                        st.session_state[sel_key].remove(iid)

                    img = item.get("imagem_url", "")
                    img_html = (f'<img src="{img}" style="width:28px;height:28px;'
                                f'object-fit:cover;border-radius:4px;border:1px solid #30363d">'
                                if img else "📦")
                    meta = " · ".join(filter(None, [
                        f"Marca: {item.get('marca','')}" if item.get('marca') else "",
                        f"SKU: {item.get('sku','')}"     if item.get('sku')   else "",
                        f"Forn: {fmc.get(item.get('fornecedor_id'),'')}" if item.get('fornecedor_id') else ""
                    ]))
                    cols[1].markdown(
                        f"{img_html} <span style='font-weight:600;font-size:13px'>"
                        f"{item.get('produto','')}</span><br>"
                        f"<span class='text-muted' style='font-size:11px'>{meta}</span>",
                        unsafe_allow_html=True)
                    cols[2].markdown(f"<span class='text-muted' style='font-size:11px'>{sec['nome']}</span>",
                                     unsafe_allow_html=True)
                    cols[3].markdown(f"<span style='font-size:12px'>{item.get('qtd','')} {item.get('unidade','')}</span>",
                                     unsafe_allow_html=True)
                    cols[4].markdown(f"<span class='text-muted' style='font-size:11px'>"
                                     f"{fmt_brl(item.get('preco_unit',0))}</span>", unsafe_allow_html=True)
                    cols[5].markdown(f"<b style='color:{info['cor']};font-size:13px'>"
                                     f"{fmt_brl(item.get('total',0))}</b>", unsafe_allow_html=True)

                    d = days_until(item.get("dt_necessidade"))
                    urg = (f"<span class='badge' style='background:rgba(248,81,73,.25);"
                           f"color:#F85149'>Atraso {abs(d)}d</span>" if d and d < 0 else "")
                    cols[6].markdown(f"{badge(item.get('status',''))} {urg}", unsafe_allow_html=True)
                    cols[7].markdown(badge(item.get("prioridade","")), unsafe_allow_html=True)

                    # Ações via popover
                    with cols[8]:
                        with st.popover("⚙️"):
                            if is_op():
                                st.caption("Acesso restrito.")
                            else:
                                cur = item.get("status", "Pendente")
                                for si, sv in enumerate(STATUS_ALL):
                                    tp  = "primary" if sv == cur else "secondary"
                                    lbl = f"✓ {sv}" if sv == cur else sv
                                    if st.button(lbl, key=f"st_{sv}_{iid}_{idx}",
                                                 use_container_width=True, type=tp):
                                        update_item(iid, {"status": sv})
                                        st.rerun()
                                st.divider()
                                if st.button("✏️ Editar", key=f"edit_btn_{iid}", use_container_width=True):
                                    st.session_state[f"ed_{iid}"] = True
                                    st.rerun()
                                # Excluir com confirmação
                                if st.session_state.get(f"conf_del_{iid}"):
                                    cd1, cd2 = st.columns(2)
                                    if cd1.button("✅ Confirmar", key=f"del_yes_{iid}",
                                                  use_container_width=True, type="primary"):
                                        delete_item(iid); st.rerun()
                                    if cd2.button("❌ Cancelar", key=f"del_no_{iid}",
                                                  use_container_width=True):
                                        st.session_state[f"conf_del_{iid}"] = False; st.rerun()
                                else:
                                    if st.button("🗑️ Excluir", key=f"del_btn_{iid}", use_container_width=True):
                                        st.session_state[f"conf_del_{iid}"] = True; st.rerun()

                    # Painel de edição — largura total da tela
                    if st.session_state.get(f"ed_{iid}"):
                        st.markdown("<div class='edit-panel'>", unsafe_allow_html=True)
                        st.markdown(f"<div style='font-weight:600;font-size:13px;"
                                    f"color:#f0f6fc;margin-bottom:8px'>"
                                    f"✏️ {item.get('produto','')}</div>", unsafe_allow_html=True)

                        forns2  = get_fornecedores()
                        fm3     = {f["nome"]: f["id"] for f in forns2}
                        fopts2  = ["(Nenhum)"] + list(fm3.keys())
                        fat     = "(Nenhum)"
                        if item.get("fornecedor_id"):
                            for f in forns2:
                                if f["id"] == item["fornecedor_id"]:
                                    fat = f["nome"]; break

                        with st.form(f"fedit_{iid}", border=False):
                            re1 = st.columns([2, 2, 2, 1.5, 1.5, 2])
                            ep   = re1[0].text_input("Produto", value=item.get("produto",""))
                            em2  = re1[1].text_input("Marca",   value=item.get("marca","") or "")
                            esk  = re1[2].text_input("SKU",     value=item.get("sku","") or "")
                            ee   = re1[3].text_input("EAN",     value=item.get("ean","") or "")
                            ef2  = re1[4].selectbox("Forn.", fopts2,
                                                    index=fopts2.index(fat) if fat in fopts2 else 0)
                            ei   = re1[5].text_input("Img URL", value=item.get("imagem_url","") or "")

                            re2  = st.columns([1, 1.5, 1.5, 1.5, 1.5, 1.5])
                            eq   = re2[0].number_input("Qtd",   min_value=0.0,
                                                        value=float(item.get("qtd",0)), step=1.0)
                            eun  = re2[1].selectbox("Unid", UNID,
                                                     index=UNID.index(item.get("unidade","UN"))
                                                     if item.get("unidade","UN") in UNID else 0)
                            epr  = re2[2].number_input("Preço", min_value=0.0,
                                                        value=float(item.get("preco_unit",0)),
                                                        step=0.01, format="%.2f")
                            eprio = re2[3].selectbox("Prio", PRIO,
                                                      index=PRIO.index(item.get("prioridade","Media"))
                                                      if item.get("prioridade") in PRIO else 1)
                            _dt_val = None
                            try:
                                if item.get("dt_necessidade"):
                                    _dt_val = date.fromisoformat(str(item["dt_necessidade"]))
                            except: pass
                            # Data em formato brasileiro — CORRIGIDO
                            edt  = re2[4].date_input("Data", value=_dt_val, format="DD/MM/YYYY")
                            eobs = re2[5].text_input("Obs", value=item.get("obs","") or "")

                            btn1, btn2 = st.columns(2)
                            saved  = btn1.form_submit_button("💾 Salvar",
                                                             type="primary", use_container_width=True)
                            cancel = btn2.form_submit_button("✖ Cancelar", use_container_width=True)

                            if cancel:
                                st.session_state[f"ed_{iid}"] = False; st.rerun()
                            if saved:
                                err = update_item(iid, {
                                    "produto": ep, "marca": em2, "sku": esk, "ean": ee,
                                    "fornecedor_id": fm3.get(ef2) if ef2 != "(Nenhum)" else None,
                                    "imagem_url": ei or None,
                                    "qtd": eq, "unidade": eun, "preco_unit": epr,
                                    "total": round(eq * epr, 2), "prioridade": eprio,
                                    "dt_necessidade": str(edt) if edt else None,
                                    "obs": eobs
                                })
                                if err:
                                    st.error(err)
                                else:
                                    st.session_state[f"ed_{iid}"] = False; st.rerun()
                        st.markdown("</div>", unsafe_allow_html=True)
            else:
                if not (busca or fst != "Todos" or fpr != "Todas"):
                    st.markdown("<span class='text-muted' style='font-size:12px'>"
                                "Nenhum item nesta seção.</span>", unsafe_allow_html=True)

    st.markdown(
        f"<div class='total-bar'>"
        f"<span class='text-muted' style='font-size:12px'>{info['nome']} — Total:</span>"
        f" <span style='color:{info['cor']};font-size:16px;font-weight:700;margin-left:6px'>"
        f"{fmt_brl(total_loja)}</span></div>",
        unsafe_allow_html=True)


def pagina_historico():
    st.markdown("<h2 style='font-size:20px;font-weight:700;color:#f0f6fc;margin-bottom:12px'>"
                "📅 Histórico</h2>", unsafe_allow_html=True)
    f1, f2, f3, f4 = st.columns(4)
    fl = f1.selectbox("Loja",       ["Todas","Distribuidora","Sublimação"], label_visibility="collapsed")
    fs = f2.selectbox("Status",     ["Todos"]+list(STATUS_HI),             label_visibility="collapsed")
    fp = f3.selectbox("Prioridade", ["Todas"]+PRIO,                        label_visibility="collapsed")
    fb = f4.text_input("",          placeholder="Buscar...",               label_visibility="collapsed")

    @st.cache_data(ttl=60, show_spinner=False)
    def load_historico():
        try:
            return sb.table("pc_itens").select("*, pc_secoes(nome,loja)") \
                      .in_("status", list(STATUS_HI)).execute().data or []
        except Exception as e:
            logger.error(f"Erro histórico: {e}"); return []

    with st.spinner("Carregando..."):
        todos = load_historico()

    if not todos:
        st.info("Nenhum item no histórico."); return

    df = pd.DataFrame(todos)
    df["loja"]      = df["pc_secoes"].apply(lambda x: x["loja"] if x else "")
    df["secao"]     = df["pc_secoes"].apply(lambda x: x["nome"] if x else "")
    df["total"]     = pd.to_numeric(df.get("total", 0), errors="coerce").fillna(0)
    df["fornecedor"]= df["fornecedor_id"].map({f["id"]:f["nome"] for f in get_fornecedores()}).fillna("")

    if fl != "Todas":
        df = df[df["loja"] == ("distribuidora" if fl == "Distribuidora" else "sublimacao")]
    if fs != "Todos": df = df[df["status"]     == fs]
    if fp != "Todas": df = df[df["prioridade"] == fp]
    if fb:            df = df[df["produto"].str.lower().str.contains(fb.lower(), na=False)]

    st.markdown(
        f"<div class='text-muted' style='font-size:12px;margin:6px 0'>"
        f"{len(df)} itens · Total: <b style='color:#58A6FF'>{fmt_brl(df['total'].sum())}</b>"
        f"</div>", unsafe_allow_html=True)
    st.dataframe(
        df[["produto","marca","sku","secao","loja","fornecedor",
            "qtd","unidade","total","prioridade","status","dt_necessidade"]]
          .rename(columns={"produto":"Produto","marca":"Marca","sku":"SKU",
                           "secao":"Seção","loja":"Loja","fornecedor":"Fornecedor",
                           "qtd":"Qtd","unidade":"Unid","total":"Total",
                           "prioridade":"Prioridade","status":"Status",
                           "dt_necessidade":"Data"}),
        use_container_width=True, hide_index=True, height=400)


def pagina_exportar():
    st.markdown("<h2 style='font-size:20px;font-weight:700;color:#f0f6fc;margin-bottom:12px'>"
                "📥 Exportar</h2>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    le  = c1.selectbox("Loja", ["Ambas","Distribuidora","Sublimação"])
    inc = c2.checkbox("Incluir Histórico")

    # CORRIGIDO: "Sublimação" com acento (estava "Sublimacao" → KeyError)
    lk_map = {"Ambas":"ambas","Distribuidora":"distribuidora","Sublimação":"sublimacao"}
    lk = lk_map.get(le, "ambas")

    if st.button("Gerar Excel", use_container_width=True, type="primary"):
        with st.spinner("Gerando..."):
            sf     = None if inc else STATUS_AT
            lojas_ = ["distribuidora","sublimacao"] if lk == "ambas" else [lk]
            wb = Workbook(); ws = wb.active; ws.title = "Compras"
            row = 1

            ws.merge_cells(f"A{row}:L{row}")
            c = ws.cell(row=row, column=1, value="GRUPO PRATES - GUIA DE COMPRAS")
            c.font      = Font(bold=True, size=14, color="FFFFFF")
            c.fill      = PatternFill("solid", start_color="0D1117")
            c.alignment = Alignment(horizontal="center")
            ws.row_dimensions[row].height = 28
            row += 1

            hdrs = ["Seção","Produto","Marca","SKU","EAN","Fornecedor",
                    "Qtd","Unid","Preço","Total","Prioridade","Status"]
            widths = [20,28,16,14,14,18,7,6,12,12,12,12]
            for i, h in enumerate(hdrs, 1):
                ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
                c = ws.cell(row=row, column=i, value=h)
                c.font      = Font(bold=True, color="FFFFFF", size=9)
                c.fill      = PatternFill("solid", start_color="21262D")
                c.alignment = Alignment(horizontal="center")
            row += 1

            fme = {f["id"]: f["nome"] for f in get_fornecedores()}
            for loja in lojas_:
                info = LOJAS[loja]
                ws.merge_cells(f"A{row}:L{row}")
                c = ws.cell(row=row, column=1, value=f"  {info['nome'].upper()}")
                c.font      = Font(bold=True, size=11, color="FFFFFF")
                c.fill      = PatternFill("solid", start_color="161B22")
                c.alignment = Alignment(horizontal="left")
                ws.row_dimensions[row].height = 20
                row += 1
                tl = 0

                for sec in get_secoes(loja):
                    for ri, item in enumerate(get_itens_secao(sec["id"], sf)):
                        zb   = "0D1117" if ri % 2 == 0 else "161B22"
                        vals = [sec["nome"], item.get("produto",""), item.get("marca",""),
                                item.get("sku",""), item.get("ean",""),
                                fme.get(item.get("fornecedor_id"),""),
                                item.get("qtd",""), item.get("unidade",""),
                                item.get("preco_unit",""), item.get("total",""),
                                item.get("prioridade",""), item.get("status","")]
                        for ci, v in enumerate(vals, 1):
                            c = ws.cell(row=row, column=ci, value=v)
                            c.font      = Font(size=9, color="E6EDF3")
                            c.fill      = PatternFill("solid", start_color=zb)
                            c.alignment = Alignment(horizontal="center" if ci > 5 else "left")
                            if ci in (9,10) and v:
                                c.number_format = '"R$" #,##0.00'
                        ws.row_dimensions[row].height = 16
                        tl  += float(item.get("total") or 0)
                        row += 1

                ws.merge_cells(f"A{row}:I{row}")
                c = ws.cell(row=row, column=1, value=f"TOTAL {info['nome'].upper()}")
                c.font      = Font(bold=True, size=10, color="FFFFFF")
                c.fill      = PatternFill("solid", start_color="0D1117")
                c.alignment = Alignment(horizontal="right")
                c = ws.cell(row=row, column=10, value=tl)
                c.font          = Font(bold=True, size=10, color="58A6FF")
                c.fill          = PatternFill("solid", start_color="0D1117")
                c.number_format = '"R$" #,##0.00'
                ws.row_dimensions[row].height = 22
                row += 2

            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            st.download_button(
                "⬇️ Baixar Excel", buf,
                file_name=f"Compras_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)


def pagina_fornecedores():
    st.markdown("<h2 style='font-size:20px;font-weight:700;color:#f0f6fc;margin-bottom:12px'>"
                "🏭 Fornecedores</h2>", unsafe_allow_html=True)
    tab_l, tab_n = st.tabs(["Lista", "Novo Fornecedor"])

    with tab_n:
        with st.form("fnovoforn"):
            na,nb,nc,nd,ne,nf = st.columns([2.5,2,1.8,2,1.8,1])
            fn   = na.text_input("Nome *")
            fc   = nb.text_input("Contato")
            ft   = nc.text_input("Telefone")
            fe   = nd.text_input("Email")
            fcnpj= ne.text_input("CNPJ")
            nf.markdown("<br>", unsafe_allow_html=True)
            if nf.form_submit_button("+ Adicionar", type="primary", use_container_width=True):
                if fn.strip():
                    create_fornecedor({"nome":fn.strip(),"contato":fc,"telefone":ft,
                                       "email":fe,"cnpj":fcnpj,"observacoes":"","ativo":True})
                    st.success("Cadastrado!"); st.rerun()
                else:
                    st.warning("Informe o nome.")

    with tab_l:
        bf    = st.text_input("", placeholder="Buscar...", label_visibility="collapsed", key="bff")
        forns = [f for f in get_fornecedores() if not bf or bf.lower() in f["nome"].lower()]
        for forn in forns:
            with st.form(f"ef_{forn['id']}"):
                fa,fb2,fc2,fd,fe2,fg,fh = st.columns([2.5,1.8,1.8,2,1.8,.7,.7])
                en   = fa.text_input("Nome",     value=forn.get("nome",""))
                ec   = fb2.text_input("Contato", value=forn.get("contato","") or "")
                et   = fc2.text_input("Telefone",value=forn.get("telefone","") or "")
                ee   = fd.text_input("Email",    value=forn.get("email","") or "")
                ecnpj= fe2.text_input("CNPJ",   value=forn.get("cnpj","") or "")
                eobs = st.text_input("Obs",      value=forn.get("observacoes","") or "",
                                     label_visibility="collapsed", placeholder="Obs...")
                fg.markdown("<br>", unsafe_allow_html=True)
                fh.markdown("<br>", unsafe_allow_html=True)
                if fg.form_submit_button("💾", use_container_width=True, type="primary"):
                    update_fornecedor(forn["id"], {"nome":en,"contato":ec,"telefone":et,
                                                   "email":ee,"cnpj":ecnpj,"observacoes":eobs})
                    st.rerun()
                if fh.form_submit_button("🗑", use_container_width=True):
                    delete_fornecedor(forn["id"]); st.rerun()


def pagina_admin():
    if u["acesso"] != "admin":
        st.error("Acesso restrito."); return
    st.markdown("<h2 style='font-size:20px;font-weight:700;color:#f0f6fc;margin-bottom:12px'>"
                "⚙️ Admin</h2>", unsafe_allow_html=True)
    tab_u, tab_s = st.tabs(["Usuários", "Seções"])

    ACESSO_MAP = {
        "op_dist":"Operador Dist.", "op_sub":"Operador Sub.", "op_ambas":"Operador Ambas",
        "distribuidora":"Gestor Dist.", "sublimacao":"Gestor Sub.",
        "ambas":"Gestor Ambas", "admin":"Administrador"
    }
    ACESSO_INV = {v: k for k, v in ACESSO_MAP.items()}

    with tab_u:
        st.markdown("#### Novo Usuário")
        with st.form("fnu"):
            c1,c2,c3,c4 = st.columns(4)
            nome   = c1.text_input("Nome")
            email  = c2.text_input("Email")
            senha  = c3.text_input("Senha", type="password")
            acesso = c4.selectbox("Acesso", list(ACESSO_MAP.keys()),
                                  format_func=lambda x: ACESSO_MAP[x])
            if st.form_submit_button("Criar", type="primary"):
                if nome and email and senha:
                    err = create_usuario(nome, email, senha, acesso)
                    if err: st.error(err)
                    else: st.success("Criado!")
                    st.rerun()

        st.markdown("#### Usuários")
        for usu in get_usuarios():
            st.markdown(
                f"<div style='background:#161b22;border:1px solid #30363d;border-radius:6px;"
                f"padding:8px 10px;margin-bottom:4px'>"
                f"<b style='color:#f0f6fc;font-size:13px'>{usu['nome']}</b>"
                f" <span class='text-muted' style='font-size:12px'>· {usu['email']} · "
                f"<b style='color:#3FB950'>{ACESSO_MAP.get(usu['acesso'], usu['acesso'])}</b>"
                f" · {'✅' if usu['ativo'] else '❌'}</span></div>",
                unsafe_allow_html=True)
            with st.form(f"eu_{usu['id']}"):
                e1, e2 = st.columns(2)
                en = e1.text_input("Nome",  value=usu["nome"])
                ee = e2.text_input("Email", value=usu["email"])
                e3, e4 = st.columns(2)
                opcoes_display = list(ACESSO_MAP.values())
                opcoes_valor   = list(ACESSO_MAP.keys())
                idx_atual = opcoes_valor.index(usu["acesso"]) if usu["acesso"] in opcoes_valor else 0
                ea_display = e3.selectbox("Acesso", opcoes_display, index=idx_atual)
                ep = e4.text_input("Nova Senha", type="password")
                s1, s2 = st.columns(2)
                if s1.form_submit_button("Salvar", type="primary"):
                    ea_valor = opcoes_valor[opcoes_display.index(ea_display)]
                    d = {"nome": en, "email": ee, "acesso": ea_valor}
                    if ep: d["senha_hash"] = hash_pwd(ep)
                    err = update_usuario(usu["id"], d)
                    if err: st.error(err)
                    else: st.success("Salvo!")
                    st.rerun()
                if s2.form_submit_button("Desativar" if usu["ativo"] else "Ativar"):
                    update_usuario(usu["id"], {"ativo": not usu["ativo"]}); st.rerun()

    with tab_s:
        for loja, info in LOJAS.items():
            st.markdown(f"#### {info['icone']} {info['nome']}")
            for sec in get_secoes(loja):
                with st.form(f"as_{sec['id']}"):
                    s1, s2, s3 = st.columns([4, 1.5, 1])
                    nn = s1.text_input("", value=sec["nome"], label_visibility="collapsed")
                    if s2.form_submit_button("Salvar"):
                        update_secao(sec["id"], nn); st.rerun()
                    if s3.form_submit_button("Excluir"):
                        delete_secao(sec["id"]); st.rerun()
            st.divider()

# ─────────────────────────────────────────────────────────────────────────────
# ROUTER
# ─────────────────────────────────────────────────────────────────────────────
pg = st.session_state.pagina
if   pg == "dashboard":     pagina_dashboard()
elif pg == "distribuidora":
    if check_perm("distribuidora"): pagina_loja("distribuidora")
    else: st.error("Acesso negado.")
elif pg == "sublimacao":
    if check_perm("sublimacao"): pagina_loja("sublimacao")
    else: st.error("Acesso negado.")
elif pg == "historico":     pagina_historico()
elif pg == "exportar":      pagina_exportar()
elif pg == "fornecedores":  pagina_fornecedores()
elif pg == "admin":
    if u["acesso"] == "admin": pagina_admin()
    else: st.error("Acesso restrito.")
